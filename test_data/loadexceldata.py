import os

import openpyxl

class LoadExcelData:

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        # book = openpyxl.load_workbook("C:\\Users\\SRIDB\\PycharmProjects\pythonProject\\test_data\\input_testdata.xlsx")
        book = openpyxl.load_workbook(os.getcwd()+'\\test_data\\input_testdata.xlsx')
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="srid
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]